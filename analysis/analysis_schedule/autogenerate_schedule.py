import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from database import SessionLocal, engine
from models.group import *
from models.lesson import *
from models.level import *
from models.student import *
from models.teacher import *
from datetime import datetime, time
import json
import re


# Функция для загрузки данных из Excel и создания словаря schedule_dict


def load_schedule_data(file_path):
    workbook = load_workbook(file_path)

    # Удаляем лишние листы
    for sheet in workbook.sheetnames:
        if 'расписание на английский оффлай' not in sheet:
            del workbook[sheet]

    workbook.save(file_path)

    # Чтение данных в DataFrame
    df = pd.read_excel(file_path)

    # Получаем уникальные группы
    groups = pd.unique(df.drop(columns=['Times', 'Teachers']).values.ravel())
    groups = [group for group in groups if pd.notna(
        group) and group.strip() != '']

    # Словарь для уровней
    levels_dict = {
        "Beginner": ["бегинер", "begginer", "beginer", "с нуля"],
        "Beginner+": ["близке к элементари", "элементари близкие к pre", "близкие к pre", 'begginer+'],
        "Elementary": ["элементари", "elementari", "элементар", 'elementary'],
        "Elementary+": ["elementari+", "элементари+", "elementari+", 'elementary+'],
        "Pre-Intermediate": ["Pre intermediate", "Preintermediate", "Pre-intermediate", "близкие к pre", 'pre-intermediate', "preintermediate"],
        "Intermediate": ["Intermediate", 'intermediate', 'intermediate'],
        "Upper-Intermediate": ["Upper Intermediate", "UpperIntermediate", 'UpperIntermediate', "upperIntermediate", "UpperIntermediate", 'upperintermediate'],
        "Advanced": ["C1"]
    }

    # Функция для нормализации уровня
    def normalize_level(level_name):
        for normalized_level, variants in levels_dict.items():
            if level_name in variants:
                return normalized_level
        return level_name

    # Словарь для расписания
    schedule_dict = {}

# Парсинг данных из DataFrame
    for group in groups:  # Перебираем каждую группу
        matches = df.iloc[:, 1:-1].isin([group])  # Находим совпадения
        filtered_matches = matches.loc[matches.any(axis=1), matches.any(axis=0)]
        for row_idx in filtered_matches.index:  # Перебираем строки
            for col_idx in filtered_matches.columns:  # Перебираем столбцы
                if filtered_matches.loc[row_idx, col_idx]:  # Если группа найдена
                    teacher = df.loc[row_idx, 'Teachers']  # Получаем преподавателя
                    time = df.loc[row_idx, 'Times']  # Получаем время
                    day = col_idx  # Получаем день недели
                    # Пропускаем NaN значения
                    if pd.isna(teacher) or pd.isna(time) or pd.isna(day):
                        continue
                    # Разбираем название группы
                    split_group = group.split('(')
                    split_group = [
                        part.replace(')', '').strip().rstrip(',')  # Удаляем только скобки и запятые
                        for part in split_group if part.strip()
                    ]
                    # Нормализуем данные
                    names = ', '.join([name.strip() for name in split_group[0].split(',')])  # Форматируем имена
                    age = [int(x) for x in re.findall(r'\d+', split_group[1])]  # Преобразуем возраст
                    level = normalize_level(split_group[2])  # Нормализуем уровень
                    # Добавляем преподавателя и группу
                    if teacher not in schedule_dict:
                        schedule_dict[teacher] = {}
                    # Добавляем день и время
                    if group in schedule_dict[teacher]:  # Если группа уже существует
                        if day not in schedule_dict[teacher][group]['Дни']:  # Если день ещё не добавлен
                            schedule_dict[teacher][group]['Дни'][day] = []
                        if time not in schedule_dict[t eacher][group]['Дни'][day]:  # Если время ещё не добавлено
                            schedule_dict[teacher][group]['Дни'][day].append(time)
                    else:  # Если группы ещё нет, создаём новую запись
                        schedule_dict[teacher][group] = {
                            'Имена': names,
                            'Возраст': age,
                            'Уровень': [level],
                            'Количество': len(names.split(',')),
                            'Дни': {day: [time]},  # Начинаем с текущего дня и времени
                        }

    return schedule_dict


file_path = 'analysis/analysis_schedule/Английский.xlsx'
# print(load_schedule_data(file_path))

# {'Айгуль': {'(Бексултан, Малика) (8) (begginer)': {'Имена': 'Бексултан,Малика', 'Возраст': '8', 'Уровень': ['Beginner'], 'Количество': 2, 'Дни': {'Monday': ['09:00-09:55'], 'Wednesday': ['09:00-09:55'], 'Friday': ['09:00-09:55']}}, '(Айсана, трое из одной семьи) (6-9лет) (бегинер)': {'Имена': 'Айсана,троеизоднойсемьи', 'Возраст': '6-9лет', 'Уровень': ['Beginner'], 'Количество': 2, 'Дни': {'Tuesday': ['09:00-09:55'], 'Thursday': ['09:00-09:55']}}, '(Арлан,Асанали) (10-11лет) (элементари)': {'Имена': 'Арлан,Асанали', 'Возраст': '10-11лет', 'Уровень': ['Elementary'], 'Количество': 2, 'Дни': {'Tuesday': ['10:00-10:55'], 'Thursday': ['10:00-10:55']}}, '(Азиза,Ислам,Мухамедияр) (10) (элементари)': {'Имена': 'Азиза,Ислам,Мухамедияр', 'Возраст': '10', 'Уровень': ['Elementary'], 'Количество': 3, 'Дни': {'Monday': ['11:10-12:05'], 'Wednesday': ['11:10-12:05'], 'Friday': ['11:10-12:05']}}, '(Ергали) (9) (begginer+)': {'Имена': 'Ергали', 'Возраст': '9', 'Уровень': ['Beginner+'], 'Количество': 1, 'Дни': {'Monday': ['12:00-12:55'], 'Wednesday': ['12:00-12:55'], 'Friday': ['12:00-12:55']}}, '(Amira) (13) (pre-intermediate+)': {'Имена': 'Amira', 'Возраст': '13', 'Уровень': ['pre-intermediate+'], 'Количество': 1, 'Дни': {'Tuesday': ['12:00-12:55'], 'Thursday': ['12:00-12:55']}}}, 'Aliza': {'(Акерке,Ислам,Аманат) (14-16лет)  (pre-inter+)': {'Имена': 'Акерке,Ислам,Аманат', 'Возраст': '14-16лет', 'Уровень': ['pre-inter+'], 'Количество': 3, 'Дни': {'Saturday': ['11:00-11:55']}}, '(Speaking club) (12-60) (pre-intermediate)': {'Имена': 'Speakingclub', 'Возраст': '12-60', 'Уровень': ['Pre-Intermediate'], 'Количество': 1, 'Дни': {'Saturday': ['13:00-13:55']}}, '(Алина,Дамеля) (10,11) (элементари)': {'Имена': 'Алина,Дамеля', 'Возраст': '10,11', 'Уровень': ['Elementary'], 'Количество': 2, 'Дни': {'Monday': ['14:00-14:55'], 'Wednesday': ['14:00-14:55']}}, '(Алина, Дамеля) (10,11) (элементари)': {'Имена': 'Алина,Дамеля', 'Возраст': '10,11', 'Уровень': ['Elementary'], 'Количество': 2, 'Дни': {'Friday': ['14:00-14:55']}}, '(Дулат, Занғар, Мирас), (13,15) (элементари +),': {'Имена': 'Дулат,Занғар,Мирас,', 'Возраст': '13,15', 'Уровень': ['Elementary+'], 'Количество': 4, 'Дни': {'Monday': ['15:00-15:55']}}, '(Нурислам, Аяулым) (13-15) (begginer) ': {'Имена': 'Нурислам,Аяулым', 'Возраст': '13-15', 'Уровень': ['Beginner'], 'Количество': 2, 'Дни': {'Tuesday': ['15:00-15:55'], 'Thursday': ['15:00-15:55']}}, '(Дулат, Занғар, Мирас), (13,15) (элементари),': {'Имена': 'Дулат,Занғар,Мирас,', 'Возраст': '13,15', 'Уровень': ['Elementary'], 'Количество': 4, 'Дни': {'Wednesday': ['15:00-15:55']}}, '(Дулат, Занғар,Мирас), (13,15) (элементари),': {'Имена': 'Дулат,Занғар,Мирас,', 'Возраст': '13,15', 'Уровень': ['Elementary'], 'Количество': 4, 'Дни': {'Friday': ['15:00-15:55']}}, '(Мұхаммеджан, Ернар,Айлана, Адина ) (12-17) (elementary) новая': {'Имена': 'Мұхаммеджан,Ернар,Айлана,Адина', 'Возраст': '12-17', 'Уровень': ['elementaryновая'], 'Количество': 4, 'Дни': {'Monday': ['16:00-16:55']}}, 'Зере (10) (бегинер)': {'Имена': 'Зере', 'Возраст': '10', 'Уровень': ['Beginner'], 'Количество':
# 1, 'Дни': {'Tuesday': ['16:00-16:55'], 'Thursday': ['16:00-16:55']}}, '(Мұхаммеджан, Ернар,Айлана, Адина ) (12-17) (elementary)': {'Имена': 'Мұхаммеджан,Ернар,Айлана,Адина', 'Возраст': '12-17', 'Уровень': ['Elementary'], 'Количество': 4, 'Дни': {'Wednesday': ['16:00-16:55'], 'Friday': ['16:00-16:55']}}, '(Акерке,Ислам,Аманат) (14-16лет)  (elementary+)': {'Имена': 'Акерке,Ислам,Аманат', 'Возраст': '14-16лет', 'Уровень': ['Elementary+'], 'Количество': 3, 'Дни': {'Tuesday': ['17:00-17:55'], 'Thursday': ['17:00-17:55']}}, '(Гульмира, Толеубек)
# (45+) (elementary)': {'Имена': 'Гульмира,Толеубек', 'Возраст': '45+', 'Уровень': ['Elementary'], 'Количество': 2, 'Дни': {'Monday': ['18:00-18:55'], 'Wednesday': ['18:00-18:55'], 'Friday': ['18:00-18:55']}}}, 'Edrich': {'(Speaking club) (12-60) (pre-intermediate)': {'Имена': 'Speakingclub', 'Возраст': '12-60', 'Уровень': ['Pre-Intermediate'], 'Количество': 1, 'Дни': {'Saturday': ['12:00-12:55']}}, '(Tore,Mustafa) (12,14) (elementary+)': {'Имена': 'Tore,Mustafa', 'Возраст': '12,14', 'Уровень': ['Elementary+'], 'Количество': 2, 'Дни': {'Monday': ['10:00-10:55'], 'Wednesday': ['10:00-10:55'], 'Friday': ['10:00-10:55']}}, '(Sara) (13) (Upper Intermediate+)': {'Имена': 'Sara', 'Возраст': '13', 'Уровень': ['UpperIntermediate+'], 'Количество': 1, 'Дни': {'Tuesday': ['10:00-10:55'], 'Thursday': ['10:00-10:55']}}, '(Alinur,Amira) (10,12) (elementary)': {'Имена': 'Alinur,Amira', 'Возраст': '10,12', 'Уровень': ['Elementary'], 'Количество': 2, 'Дни': {'Monday': ['11:10-12:05'], 'Wednesday': ['11:10-12:05'], 'Friday': ['11:10-12:05']}}, '(Umar, Korkem,Zhanel,Barsbek) (9,11) (elementary)': {'Имена': 'Umar,Korkem,Zhanel,Barsbek', 'Возраст': '9,11', 'Уровень': ['Elementary'], 'Количество': 4, 'Дни': {'Tuesday': ['11:10-12:05'], 'Thursday': ['11:10-12:05']}}, '(Alina) (14) (intermediate) (private)': {'Имена': 'Alina', 'Возраст': '14', 'Уровень': ['Intermediate'], 'Количество': 1, 'Дни': {'Saturday': ['11:10-12:05'], 'Tuesday': ['15:00-15:55'], 'Thursday': ['15:00-15:55']}}, '(Kaysar, Aibarkhan,Aidos) (10) (pre-intermediate) ': {'Имена': 'Kaysar,Aibarkhan,Aidos', 'Возраст': '10', 'Уровень': ['Pre-Intermediate'], 'Количество': 3, 'Дни': {'Tuesday': ['12:00-12:55'], 'Thursday': ['12:00-12:55']}}, '(Speaking club) (6-12) (beginer,elementary)': {'Имена': 'Speakingclub', 'Возраст': '6-12', 'Уровень': ['beginer,elementary'], 'Количество': 1, 'Дни': {'Saturday': ['13:00-13:55']}}, '(Sanjar) (14-16) (Upper Intermediate+)': {'Имена': 'Sanjar', 'Возраст': '14-16', 'Уровень': ['UpperIntermediate+'], 'Количество': 1, 'Дни': {'Monday': ['16:00-16:55'], 'Wednesday': ['16:00-16:55'], 'Friday': ['16:00-16:55']}}, '(Arlan) (13) (elementary)': {'Имена': 'Arlan', 'Возраст': '13', 'Уровень': ['Elementary'], 'Количество': 1, 'Дни': {'Tuesday': ['16:00-16:55'], 'Thursday': ['16:00-16:55']}}, '(Azhar,Inkara, Zhanel,Adel, Adil, Amir) (16) (intermediate)': {'Имена':
# 'Azhar,Inkara,Zhanel,Adel,Adil,Amir', 'Возраст': '16', 'Уровень': ['Intermediate'], 'Количество': 6, 'Дни': {'Monday': ['17:00-17:55'], 'Wednesday': ['17:00-17:55'], 'Friday': ['17:00-17:55']}}}, 'Айжан': {'(Speaking club) (6-12) (beginer,elementary)': {'Имена': 'Speakingclub', 'Возраст': '6-12', 'Уровень': ['beginer,elementary'], 'Количество': 1, 'Дни': {'Saturday': ['13:00-13:55']}}, '(Алая,Томирис) (5) (beginer)': {'Имена': 'Алая,Томирис', 'Возраст': '5', 'Уровень': ['Beginner'], 'Количество': 2, 'Дни': {'Saturday': ['10:00-10:55'], 'Sunday': ['10:00-10:55']}}, '(Назерьке, Айсулу) (25+) (elementary +)': {'Имена': 'Назерьке,Айсулу', 'Возраст': '25+', 'Уровень': ['Elementary+'], 'Количество': 2, 'Дни': {'Saturday': ['11:00-11:55'], 'Sunday': ['11:00-11:55']}}, '(Жанали, Марьям, Сафия, Асанали) (6-7) (elementary) (private)': {'Имена': 'Жанали,Марьям,Сафия,Асанали', 'Возраст': '6-7', 'Уровень': ['Elementary'], 'Количество': 4, 'Дни': {'Saturday': ['12:00-12:55']}}, ' 12:15 (Жанали, Марьям, Сафия, Асанали) (6-7) (elementary) (private)': {'Имена': '12:15', 'Возраст': 'Жанали,Марьям,Сафия,Асанали', 'Уровень': ['6-7'], 'Количество': 1, 'Дни': {'Sunday': ['12:00-12:55']}}, '(София, Сара, София) (9лет) (beginer)': {'Имена': 'София,Сара,София', 'Возраст': '9лет', 'Уровень': ['Beginner'], 'Количество': 3, 'Дни': {'Saturday': ['14:00-14:55'], 'Sunday': ['14:00-14:55']}}}, 'Меруерт': {'(Амина, Аида,) (14-15) (elementary+)': {'Имена': 'Амина,Аида,', 'Возраст': '14-15', 'Уровень': ['Elementary+'], 'Количество': 3, 'Дни': {'Monday': ['18:00-18:55'], 'Wednesday': ['18:00-18:55'], 'Friday': ['18:00-18:55']}}, '(Амир, Дамели) (6-8) (бегинер)': {'Имена': 'Амир,Дамели', 'Возраст': '6-8', 'Уровень': ['Beginner'], 'Количество': 2, 'Дни': {'Tuesday': ['18:00-18:55'], 'Thursday': ['18:00-18:55']}}, '(Мухаммед, Абдрахман) (7-9) (beginner)': {'Имена': 'Мухаммед,Абдрахман', 'Возраст': '7-9', 'Уровень': ['beginner'], 'Количество': 2, 'Дни': {'Monday': ['19:00-19:55'], 'Wednesday': ['19:00-19:55'], 'Friday': ['19:00-19:55']}}, '(Молдир) (33) (elementary) (private)': {'Имена': 'Молдир', 'Возраст': '33', 'Уровень': ['Elementary'], 'Количество': 1, 'Дни': {'Tuesday': ['19:00-19:55'], 'Friday': ['20:00-20:55']}}}, 'Тахир ': {'(Мустафа) (12) (private)': {'Имена': 'Мустафа', 'Возраст': '12', 'Уровень': ['private'], 'Количество': 1, 'Дни': {'Tuesday': ['10:00-10:55'], 'Thursday': ['10:00-10:55']}}, '(Каракат,Мирас) (12,14) (private)': {'Имена': 'Каракат,Мирас', 'Возраст': '12,14', 'Уровень': ['private'], 'Количество': 2, 'Дни': {'Monday': ['16:00-16:55'], 'Wednesday': ['16:00-16:55']}}}}


# def save_students_to_db(data):
#     session=SessionLocal()
#     for teacher, group in data.items():
#         # print(f"{teacher} - учитель {group} - его группы")
#         for group, detail in group.items():
#             names = detail['Имена'].strip().split(',')
#             # age=min(list(map(int,re.findall(r'\d+',detail['Возраст']))))
#             age=min(detail['Возраст'])
#             level_str = detail['Уровень'][0].strip().lower() if isinstance(detail['Уровень'], list) else detail['Уровень'].strip().lower()
#             english_level = next((level for level in EnglishLevel if level.value.lower() == level_str),None)
#             # if not english_level:
#             #     raise ValueError(f"Unknown English level: {level_str}")
#             lesson_days=detail['Дни']
#             lesson_times=detail['Дни']
#             age_group=AgeGroup.KIDS if 6<=age<13 else AgeGroup.TEENS if 13<=age<18 else AgeGroup.ADULTS
#             # school_shift=SchoolShift.FIRST_SHIFT if 
#             teacher_type=TeacherType.NATIVE if teacher=='Edrich' else TeacherType.LOCAL


#             for name in names:
#                 student=Student(
#                     full_name=name,
#                     age=age,
#                     english_level=english_level,
#                     lesson_days=json.dumps(lesson_days),
#                     lesson_times=json.dumps(lesson_times),
#                     age_group=age_group,
#                     teacher_type=teacher_type
#                 )
#                 session.add(student)
#     session.commit()
#     session.close()

# Функция для сравнения времени занятий
def compare_times(time1, time2):
    """
    Сравнивает два времени, даже если они указаны в разных форматах.
    Например, "18:00" и "18:00-18:55" считаются совпадающими.
    """
    def time_to_minutes(time_str):
        # Преобразует время в минуты для сравнения
        if '-' in time_str:  # Если время указано в формате "18:00-18:55"
            start, _ = time_str.split('-')  # Берем только начало времени
            return int(start.split(':')[0]) * 60 + int(start.split(':')[1])
        else:  # Если время указано в формате "18:00"
            return int(time_str.split(':')[0]) * 60 + int(time_str.split(':')[1])

    # Сравниваем время в минутах
    return time_to_minutes(time1) == time_to_minutes(time2)

