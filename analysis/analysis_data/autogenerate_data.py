import pandas as pd
from openpyxl import load_workbook
import re

# Функция для загрузки данных из Excel и создания словаря schedule_dict

file_path = 'analysis/analysis_data/Продажи.xlsx'
workbook = load_workbook(file_path)

# Удаляем лишние листы
for sheet in workbook.sheetnames:
    if 'Оплата Февраль' not in sheet or 'копия' in sheet:
        del workbook[sheet]
workbook.save(file_path)

# Чтение данных в DataFrame
df = pd.read_excel(file_path)

# Очистка данных: удаляем строки, где все значения NaN
df_cleaned = df.dropna(how='all')

# Создаем пустой словарь для хранения данных
full_student_data = {}

# Функция для нормализации дней недели
def normalize_days(day_string):
    # Словарь для перевода дней недели на английский
    day_synonyms = {
        # Monday
        "пон": "Monday", "поне": "Monday", "пн": "Monday", "понед": "Monday", "понедельн": "Monday", "понедельник": "Monday", "пнд": "Monday",
        # Tuesday
        "вт": "Tuesday", "вто": "Tuesday", "втр": "Tuesday", "втор": "Tuesday", "вторн": "Tuesday", "втоник": "Tuesday", "вторник": "Tuesday",
        # Wednesday
        "ср": "Wednesday", "сре": "Wednesday", "сред": "Wednesday", "среда": "Wednesday",
        # Thursday
        "чет": "Thursday", "четв": "Thursday", "чт": "Thursday", "чевтерг": "Thursday", "четвер": "Thursday", "четверг": "Thursday",
        # Friday
        "пят": "Friday", "пятн": "Friday", "пт": "Friday", "пятниц": "Friday", "пятница": "Friday",
        # Saturday
        "суб": "Saturday", "субб": "Saturday", "сб": "Saturday", "субота": "Saturday", "суббота": "Saturday",
        # Sunday
        "вос": "Sunday", "воскр": "Sunday", "вс": "Sunday", "воскрес": "Sunday", "воскресенье": "Sunday"
    }
    # Если входная строка пустая или None, возвращаем пустой список
    if not isinstance(day_string, str) or not day_string.strip():
        return []
    # Разделяем строку по любым разделителям: запятые, пробелы, тире, слэши
    raw_days = re.split(r'[,\s\-/]+', day_string.strip())
    normalized_days = []
    for day in raw_days:
        day = day.strip().lower()  # Удаляем пробелы и приводим к нижнему регистру
        if day in day_synonyms:
            normalized_days.append(day_synonyms[day])  # Нормализуем день
        elif day:
            continue  # Если день не найден, добавляем "НЕТУ"
    return normalized_days

# Функция для извлечения времени
def extract_time(schedule_string):
    time_pattern = r'(?:в\s+)?(\d{1,2}[:.-]\d{2})'
    matches = re.findall(time_pattern, schedule_string) if isinstance(schedule_string, str) else []
    return [time.replace('.', ':').replace('-', ':') for time in matches]

# Функция для извлечения имени родителя
parent_keywords = ["мама", "папа", "бабушка", "дедушка",'дядя','тетя','брат','сестра',]
# Функция для извлечения имени родителя с указанием роли (мама, папа и т.д.)
def extract_parent_name(fio, phone):
    # Очищаем строку от лишних символов (скобок)
    fio = re.sub(r'[()]', '', fio).strip()
    
    # Ищем ключевые слова в строке "ФИО"
    for keyword in parent_keywords:
        if keyword in fio.lower():
            # Разделяем строку по ключевому слову
            parts = fio.lower().split(keyword)
            if len(parts) > 1:
                # Берем часть после ключевого слова
                parent_part = parts[1].strip()
                
                # Извлекаем первое или два слова как имя родителя
                parent_words = parent_part.split()[:2]  # Берем максимум два слова
                
                # Приводим каждое слово к формату с заглавной буквы
                parent_name = " ".join(word.capitalize() for word in parent_words)
                
                # Проверяем, совпадает ли имя родителя с именем в телефоне
                if phone and parent_name.lower() in phone.lower():
                    return f"{keyword.capitalize()}: {parent_name}"
                
                # Если нет совпадения, проверяем, есть ли имя родителя в самом ФИО
                if parent_name.lower() in fio.lower():
                    return f"{keyword.capitalize()}: {parent_name}"
    
    # Если ключевые слова не найдены, возвращаем None
    return None

# Функция для извлечения имени студента
def extract_student_name(fio):
    # Удаляем скобки и лишние символы
    fio = re.sub(r'[()]', '', fio).strip()
    
    # Разделяем строку на части
    parts = fio.split()
    
    # Ищем ключевые слова, связанные с родителями
    for i, part in enumerate(parts):
        if part.lower() in parent_keywords:
            # Если находим ключевое слово, берем все, что до него, как имя студента
            return " ".join(parts[:i]).strip()
    
    # Если ключевых слов нет, возвращаем всю строку как имя студента
    return fio.strip()

# Проходимся по каждой строке DataFrame
for index, row in df_cleaned.iterrows():
    # Проверяем, содержит ли строка слово "расшифровка" (или другое условие для прерывания)
    if row.astype(str).str.lower().str.contains('расшифровка', na=False).any():
        break
    
    # Создаем временный словарь для текущей строки
    student_info = {}
    
    # Проходимся по каждому столбцу в строке
    for column_name, value in row.items():
        # Проверяем название столбца и выполняем соответствующие действия
        if column_name == "ФИО":
            # Если столбец называется "ФИО", используем его как ключ
            student_name = value.strip() if isinstance(value, str) else str(value)
            
            # Извлекаем имя студента
            student_info["Имя"] = extract_student_name(student_name)
            
            # Извлекаем имя родителя
            phone = row.get("Телефон", "").strip() if isinstance(row.get("Телефон", ""), str) else ""
            parent_name = extract_parent_name(student_name, phone)
            student_info["Имя родителя/родственника"] = parent_name
        
        elif column_name == "Телефон":
            # Если столбец называется "Телефон", сохраняем номер телефона
            student_info["Контакт"] = value.strip() if isinstance(value, str) else str(value)
        
        elif column_name == "Дни занятий":
            # Если столбец называется "Дни занятий", нормализуем дни недели
            days = normalize_days(value) if isinstance(value, str) else []
            times = extract_time(value) if isinstance(value, str) else []
            student_info["Дни занятий"] = days
            student_info["Время занятий"] = times
        
        elif column_name == "Сумма оплаты":
            # Если столбец называется "Сумма оплаты", преобразуем значение в число
            student_info["Ожидаемая оплата"] = int(value.replace(" ", "")) if isinstance(value, str) else value
        
        elif column_name == "Факт":
            # Если столбец называется "Факт", преобразуем значение в число
            student_info["Фактическая оплата"] = int(value.replace(" ", "")) if isinstance(value, str) else value
        
        elif column_name == "Дата":
            # Если столбец называется "Дата", сохраняем дату
            student_info["Дата оплаты"] = value.strip() if isinstance(value, str) else str(value)
        
        elif column_name == "Статус":
            # Если столбец называется "Статус", сохраняем статус
            student_info["Статус"] = value.strip() if isinstance(value, str) else str(value)
    
    # Добавляем информацию о студенте в общий словарь
    full_student_data[student_name] = student_info

# Вывод результата
# for name, info in full_student_data.items():
#     print(f"ФИО: {name}")
#     for key, value in info.items():
#         print(f"  {key}: {value}")
#     print()

print(full_student_data)